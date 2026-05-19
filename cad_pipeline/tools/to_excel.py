#!/usr/bin/env python3
"""to_excel.py — Export chat/tool JSON data to multi-sheet Excel.

Source:
  1. JSON file from chat/tool extraction (cad_llm_schema-like)

Sheets generated:
  📋 Summary       — key metrics (file info, entity counts, scale)
  🗂  Layers        — all layers with category, entity counts, filterable
  🧱 Symbols       — INSERT blocks grouped by layer category
  🚪 Equipment     — doors/windows/equipment counts
  📐 Areas         — closed polyline boundaries with area in m²
  🏠 Rooms         — room names paired with areas
  📝 Notes         — annotation texts

Usage:
  python3 to_excel.py report_payload.json
  python3 to_excel.py report_payload.json -o report.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo


# ── Design tokens ──────────────────────────────────────────────────────────
C_HEADER_BG   = "0F3460"   # deep navy
C_HEADER_FG   = "FFFFFF"
C_SUB_BG      = "16213E"   # dark navy
C_ACCENT      = "E94560"   # red-pink
C_ALT_ROW     = "EEF2FF"   # light blue-gray
C_BORDER      = "AAAACC"
C_TITLE_FG    = "0F3460"
C_SHEET_TABS  = ["0F3460","16213E","533483","2D6A4F","B5451B","4A4E69","3D405B"]

FONT_MAIN  = "Calibri"


# ── Style helpers ──────────────────────────────────────────────────────────

def _border(color: str = C_BORDER) -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _header_row(ws, row: int, values: list, widths: list | None = None,
                bg: str = C_HEADER_BG, fg: str = C_HEADER_FG) -> None:
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(name=FONT_MAIN, bold=True, color=fg, size=10)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _border()
    if widths:
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w


def _data_row(ws, row: int, values: list, alt: bool = False) -> None:
    bg = C_ALT_ROW if alt else "FFFFFF"
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(name=FONT_MAIN, size=9)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(vertical="center", wrap_text=False)
        c.border    = _border()


def _sheet_title(ws, title: str, subtitle: str = "") -> None:
    ws.cell(1, 1, title).font = Font(name=FONT_MAIN, bold=True, size=14, color=C_TITLE_FG)
    if subtitle:
        ws.cell(2, 1, subtitle).font = Font(name=FONT_MAIN, size=9, color="666666")
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 14


def _add_table(ws, ref: str, name: str, style: str = "TableStyleMedium2") -> None:
    """Add Excel auto-filter table to a range."""
    try:
        tbl = Table(displayName=name.replace(" ", "_"), ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(name=style, showRowStripes=True)
        ws.add_table(tbl)
    except Exception:
        pass  # table already exists or invalid range — skip


def _freeze_and_filter(ws, freeze: str = "A4") -> None:
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions


def _set_tab_color(ws, idx: int) -> None:
    ws.sheet_properties.tabColor = C_SHEET_TABS[idx % len(C_SHEET_TABS)]


# ── Sheet builders ─────────────────────────────────────────────────────────

def build_summary_sheet(ws, data: dict) -> None:
    ws.title = "📋 Summary"
    _set_tab_color(ws, 0)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 55

    meta = data.get("meta", {})
    stats = data.get("stats", {})

    rows = [
        ("ファイル名",       meta.get("filename", "—")),
        ("分析日時",         meta.get("date", "—")),
        ("縮尺",             meta.get("scale", "—")),
        ("",                 ""),
        ("レイヤー総数",      stats.get("total_layers", 0)),
        ("分類済みレイヤー", stats.get("classified_layers", 0)),
        ("総エンティティ",   stats.get("total_entities", 0)),
        ("",                 ""),
        ("INSERT (ブロック)", stats.get("total_inserts", 0)),
        ("TEXT/MTEXT",       stats.get("total_texts", 0)),
        ("閉合ポリライン",   stats.get("total_boundaries", 0)),
        ("HATCH",            stats.get("total_hatches", 0)),
    ]

    ws.cell(1, 1, "📋 ファイルサマリー").font = Font(name=FONT_MAIN, bold=True, size=14, color=C_TITLE_FG)
    ws.row_dimensions[1].height = 26
    ws.cell(2, 1, meta.get("title", "")).font = Font(name=FONT_MAIN, size=10, color="444444")

    _header_row(ws, 3, ["項目", "値"])

    for i, (k, v) in enumerate(rows, 4):
        alt = (i % 2 == 0)
        if k == "":
            ws.row_dimensions[i].height = 6
            continue
        _data_row(ws, i, [k, v], alt)

    ws.freeze_panes = "A4"


def build_layers_sheet(ws, layers_data: list[dict]) -> None:
    ws.title = "🗂 Layers"
    _set_tab_color(ws, 1)

    headers = ["レイヤー名", "カテゴリ", "ラベル", "総エンティティ", "INSERT", "TEXT", "LINE/LWPOLY", "HATCH", "主なタイプ"]
    widths  = [38, 14, 14, 14, 10, 10, 14, 10, 38]

    _sheet_title(ws, "🗂 レイヤー分類一覧", f"{len(layers_data)} layers")
    _header_row(ws, 3, headers, widths)

    for i, row in enumerate(layers_data, 4):
        et = row.get("entity_types", {})
        top3 = ", ".join(f"{k}:{v}" for k, v in
                         sorted(et.items(), key=lambda x: -x[1])[:3])
        vals = [
            row.get("layer", ""),
            row.get("category", "other"),
            row.get("label", "その他"),
            row.get("total", 0),
            et.get("INSERT", 0),
            et.get("TEXT", 0) + et.get("MTEXT", 0),
            et.get("LINE", 0) + et.get("LWPOLYLINE", 0),
            et.get("HATCH", 0),
            top3,
        ]
        _data_row(ws, i, vals, alt=(i % 2 == 0))

    _freeze_and_filter(ws, "A4")
    if len(layers_data) > 0:
        end_col = get_column_letter(len(headers))
        _add_table(ws, f"A3:{end_col}{3+len(layers_data)}", "LayerTable")


def build_symbols_sheet(ws, symbols_data: list[dict]) -> None:
    ws.title = "🧱 Symbols"
    _set_tab_color(ws, 2)

    headers = ["ブロック名", "カテゴリ", "ラベル", "挿入数", "出現ファイル数", "主なレイヤー"]
    widths  = [45, 14, 14, 10, 12, 45]

    _sheet_title(ws, "🧱 ブロック/シンボル一覧", f"{len(symbols_data)} unique blocks")
    _header_row(ws, 3, headers, widths)

    for i, row in enumerate(symbols_data, 4):
        top_layers = ", ".join(list(row.get("layers", {}).keys())[:3])
        vals = [
            row.get("block_name", ""),
            row.get("category", "other"),
            row.get("label", "その他"),
            row.get("count", 0),
            row.get("files", 0),
            top_layers,
        ]
        _data_row(ws, i, vals, alt=(i % 2 == 0))

        # Highlight high-count blocks
        count_cell = ws.cell(i, 4)
        if row.get("count", 0) >= 100:
            count_cell.font = Font(name=FONT_MAIN, bold=True, color=C_ACCENT, size=9)

    _freeze_and_filter(ws, "A4")
    if symbols_data:
        end_col = get_column_letter(len(headers))
        _add_table(ws, f"A3:{end_col}{3+len(symbols_data)}", "SymbolTable")


def build_equipment_sheet(ws, equip_data: dict, cat_data: dict) -> None:
    ws.title = "🚪 Equipment"
    _set_tab_color(ws, 3)

    # Section 1: By semantic (block name pattern)
    ws.cell(1, 1, "🚪 建具・設備集計").font = Font(name=FONT_MAIN, bold=True, size=14, color=C_TITLE_FG)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 12

    _header_row(ws, 3, ["種別（パターン）", "数量"])
    row = 4
    for sem, cnt in sorted(equip_data.get("by_semantic", {}).items(), key=lambda x: -x[1]):
        _data_row(ws, row, [sem, cnt], alt=(row % 2 == 0))
        row += 1

    # Section 2: By layer category (right side)
    ws.cell(1, 4, "レイヤーカテゴリ別 INSERT 集計").font = Font(name=FONT_MAIN, bold=True, size=11, color=C_SUB_BG)
    _header_row(ws, 3, ["カテゴリ", "挿入数"], bg=C_SUB_BG)

    r = 4
    for cat, entry in sorted(cat_data.items(), key=lambda x: -x[1].get("count", 0)):
        _data_row(ws, r, [entry.get("label", cat), entry.get("count", 0)], alt=(r % 2 == 0))
        # move to D/E columns
        ws.cell(r, 4).value = entry.get("label", cat)
        ws.cell(r, 5).value = entry.get("count", 0)
        for col in [4, 5]:
            c = ws.cell(r, col)
            c.font   = Font(name=FONT_MAIN, size=9)
            c.fill   = PatternFill("solid", fgColor=C_ALT_ROW if r % 2 == 0 else "FFFFFF")
            c.border = _border()
        # clear original A/B placement
        ws.cell(r, 1).value = ws.cell(r, 1).value
        r += 1

    ws.freeze_panes = "A4"


def build_areas_sheet(ws, areas_data: list[dict]) -> None:
    ws.title = "📐 Areas"
    _set_tab_color(ws, 4)

    headers = ["カテゴリ", "レイヤー", "面積 (㎡)", "境界点数"]
    widths  = [16, 35, 14, 12]

    _sheet_title(ws, "📐 閉合ポリライン境界（面積）", f"{len(areas_data)} boundaries")
    _header_row(ws, 3, headers, widths)

    for i, row in enumerate(areas_data, 4):
        vals = [
            row.get("label", row.get("category", "other")),
            row.get("layer", ""),
            round(row.get("area_m2", 0), 2),
            row.get("point_count", 0),
        ]
        _data_row(ws, i, vals, alt=(i % 2 == 0))
        # Right-align area
        ws.cell(i, 3).alignment = Alignment(horizontal="right")

    _freeze_and_filter(ws, "A4")
    if areas_data:
        end_col = get_column_letter(len(headers))
        _add_table(ws, f"A3:{end_col}{3+len(areas_data)}", "AreaTable")


def build_rooms_sheet(ws, rooms_data: list[dict]) -> None:
    ws.title = "🏠 Rooms"
    _set_tab_color(ws, 5)

    headers = ["室名", "用途", "面積 (㎡)", "レイヤー", "X", "Y"]
    widths  = [28, 22, 12, 28, 12, 12]

    _sheet_title(ws, "🏠 室名・面積一覧", f"{len(rooms_data)} rooms")
    _header_row(ws, 3, headers, widths)

    for i, row in enumerate(rooms_data, 4):
        vals = [
            row.get("name", ""),
            row.get("purpose", "その他"),
            round(row.get("area_m2", 0), 2) if row.get("area_m2") else "",
            row.get("layer", ""),
            round(row.get("x", 0), 1),
            round(row.get("y", 0), 1),
        ]
        _data_row(ws, i, vals, alt=(i % 2 == 0))

    _freeze_and_filter(ws, "A4")


def build_notes_sheet(ws, notes_data: list[dict]) -> None:
    ws.title = "📝 Notes"
    _set_tab_color(ws, 6)

    headers = ["テキスト", "カテゴリ", "レイヤー", "X", "Y"]
    widths  = [80, 14, 28, 10, 10]

    _sheet_title(ws, "📝 注記・テキスト一覧", f"{len(notes_data)} items")
    _header_row(ws, 3, headers, widths)

    for i, row in enumerate(notes_data, 4):
        vals = [
            row.get("text", ""),
            row.get("category", ""),
            row.get("layer", ""),
            round(row.get("x", 0), 1),
            round(row.get("y", 0), 1),
        ]
        _data_row(ws, i, vals, alt=(i % 2 == 0))
        ws.cell(i, 1).alignment = Alignment(wrap_text=True)

    _freeze_and_filter(ws, "A4")


# ── LLM JSON mode (from cad_llm_schema.json output) ──────────────────────

def export_llm_json_excel(json_path: Path, out_path: Path) -> None:
    """Build Excel directly from LLM output JSON (follows cad_llm_schema.json).

    This is the recommended path when data comes from an LLM:
      LLM → JSON (schema) → this function → Excel
    """
    print(f"  Loading LLM JSON: {json_path.name} …", end=" ", flush=True)
    data: dict = json.loads(json_path.read_text(encoding="utf-8"))
    print("OK")

    # Validate top-level keys
    for key in ("meta", "layers", "symbols", "equipment", "areas", "rooms", "notes"):
        if key not in data:
            print(f"  ⚠  Missing key '{key}' — will use empty list/dict")

    wb = Workbook()
    wb.remove(wb.active)

    # ── Summary ──
    meta  = data.get("meta", {})
    stats = {
        "total_layers":      len(data.get("layers", [])),
        "classified_layers": sum(1 for l in data.get("layers", []) if l.get("category") != "other"),
        "total_entities":    sum(l.get("total", 0) for l in data.get("layers", [])),
        "total_inserts":     sum(s.get("count", 0) for s in data.get("symbols", [])),
        "total_texts":       sum(l.get("text_count", 0) for l in data.get("layers", [])),
        "total_boundaries":  len(data.get("areas", [])),
        "total_hatches":     sum(l.get("hatch_count", 0) for l in data.get("layers", [])),
    }
    ws_sum = wb.create_sheet("📋 Summary")
    build_summary_sheet(ws_sum, {"meta": meta, "stats": stats})

    # ── Layers ──
    ws_lay = wb.create_sheet("🗂 Layers")
    layers_list = []
    for l in data.get("layers", []):
        layers_list.append({
            "layer":    l.get("layer", ""),
            "category": l.get("category", "other"),
            "label":    l.get("label", "その他"),
            "total":    l.get("total", 0),
            "entity_types": {
                "INSERT":     l.get("insert_count", 0),
                "TEXT":       l.get("text_count", 0),
                "LINE":       l.get("line_count", 0),
                "LWPOLYLINE": 0,
                "HATCH":      l.get("hatch_count", 0),
            },
        })
    build_layers_sheet(ws_lay, layers_list)

    # ── Symbols ──
    ws_sym = wb.create_sheet("🧱 Symbols")
    syms_list = []
    for s in data.get("symbols", []):
        syms_list.append({
            "block_name": s.get("block_name", ""),
            "category":   s.get("category", "other"),
            "label":      s.get("label", "その他"),
            "count":      s.get("count", 0),
            "files":      s.get("files", 1),
            "layers":     {s.get("layer", ""): s.get("count", 0)},
        })
    build_symbols_sheet(ws_sym, syms_list)

    # ── Equipment ──
    eq = data.get("equipment", {})
    block_summary = {
        "by_semantic": {row["type"]: row["count"] for row in eq.get("by_type", [])},
        "by_name": {},
    }
    cat_summary = {
        row["category"]: {"label": row["label"], "count": row["count"], "blocks": {}, "layers": {}}
        for row in eq.get("by_category", [])
    }
    ws_eq = wb.create_sheet("🚪 Equipment")
    build_equipment_sheet(ws_eq, block_summary, cat_summary)

    # ── Areas ──
    areas_list = []
    for a in data.get("areas", []):
        areas_list.append({
            "label":       a.get("label", a.get("category", "other")),
            "layer":       a.get("layer", ""),
            "area_m2":     a.get("area_m2", 0),
            "point_count": a.get("point_count", 0),
        })
    areas_list.sort(key=lambda x: -x["area_m2"])
    ws_area = wb.create_sheet("📐 Areas")
    build_areas_sheet(ws_area, areas_list)

    # ── Rooms ──
    ws_room = wb.create_sheet("🏠 Rooms")
    build_rooms_sheet(ws_room, data.get("rooms", []))

    # ── Notes ──
    notes_list = [
        {"text": n.get("text", ""), "category": n.get("category", ""),
         "layer": n.get("layer", ""), "x": n.get("x", 0), "y": n.get("y", 0)}
        for n in data.get("notes", [])
    ]
    ws_note = wb.create_sheet("📝 Notes")
    build_notes_sheet(ws_note, notes_list)

    wb.save(str(out_path))
    sz = out_path.stat().st_size / 1024
    print(f"  Saved: {out_path}  ({sz:.0f} KB)  — {len(wb.sheetnames)} sheets")


# ── Main ───────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Export chat/tool JSON data to multi-sheet Excel"
    )
    parser.add_argument("input", help="Input JSON from chat/tool extraction")
    parser.add_argument("-o", "--output", default=None, help="Output .xlsx path")
    args = parser.parse_args()

    inp = Path(args.input)
    if not inp.exists():
        print(f"Error: {inp} not found")
        sys.exit(1)

    out = Path(args.output) if args.output else inp.with_suffix(".xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)

    print(f"\n[to_excel] {inp.name} → {out}\n")
    if inp.suffix.lower() != ".json":
        print("Error: input must be a JSON file generated from chat/tool data.")
        sys.exit(1)
    export_llm_json_excel(inp, out)

    print("\nDone ✅")


if __name__ == "__main__":
    main()
